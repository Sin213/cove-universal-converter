#!/usr/bin/env python3
"""Smoke-test every advertised conversion route.

Discovers routes from ``cove_converter.routing.SUPPORTED_FORMATS``, generates
tiny valid sample inputs in a temp directory, drives each engine's
``_convert()`` method directly (no QThread, no GUI), and reports
PASS / FAIL / SKIP per (input_ext -> output_ext) pair.

Usage:
    python scripts/smoke_conversions.py [--engine NAME ...] [--input EXT ...]
                                        [--quiet] [--keep-temp]

Exit code is non-zero if any required (non-skipped) conversion fails.
"""

from __future__ import annotations

import argparse
import io
import json
import os
import plistlib
import shutil
import subprocess
import sys
import tarfile
import tempfile
import time
import traceback
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable
from unittest import mock

# Headless Qt — workers import PySide6 transitively via base.QThread.
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from cove_converter import routing  # noqa: E402
from cove_converter.engines import WORKER_REGISTRY  # noqa: E402


# ---- Optional-dependency probes -------------------------------------------


def _probe_module(name: str) -> str | None:
    try:
        __import__(name)
        return None
    except Exception as exc:  # pragma: no cover - probe wrapper
        return f"{name}: {exc}"


def _probe_binary(name: str) -> str | None:
    if shutil.which(name) is None:
        return f"{name} not on PATH"
    return None


# Keyed by SUPPORTED_FORMATS engine name.
ENGINE_DEPS: dict[str, list[Callable[[], str | None]]] = {
    "Pillow": [lambda: _probe_module("PIL")],
    "FFmpeg": [lambda: _probe_binary("ffmpeg")],
    "Pandoc": [lambda: _probe_binary("pandoc")],
    "Pdf": [
        lambda: _probe_module("pypdf"),
        lambda: _probe_module("xhtml2pdf"),
        lambda: _probe_binary("pandoc"),
    ],
    "Subtitle": [],
    "Spreadsheet": [lambda: _probe_module("openpyxl")],
    "Archive": [],
    "Data": [lambda: _probe_module("yaml")],
}

# Per-extension extra requirements (HEIC needs pillow-heif, etc.).
EXT_DEPS: dict[str, list[Callable[[], str | None]]] = {
    ".heic": [lambda: _probe_module("pillow_heif")],
    ".heif": [lambda: _probe_module("pillow_heif")],
}


def missing_dependencies(engine: str, in_ext: str, out_ext: str) -> list[str]:
    misses: list[str] = []
    for probe in ENGINE_DEPS.get(engine, []):
        m = probe()
        if m:
            misses.append(m)
    for ext in (in_ext, out_ext):
        for probe in EXT_DEPS.get(ext, []):
            m = probe()
            if m and m not in misses:
                misses.append(m)
    return misses


# ---- Sample generators -----------------------------------------------------
#
# Each generator writes a single tiny valid file at ``dest`` and returns the
# same path. Generators are gated on what optional libs / binaries exist; if
# generation fails the route gets SKIPPED with the reason.


def _gen_txt(dest: Path) -> Path:
    dest.write_text("Cove smoke-test sample.\nLine two.\n", encoding="utf-8")
    return dest


def _gen_md(dest: Path) -> Path:
    dest.write_text("# Smoke test\n\nA short paragraph.\n", encoding="utf-8")
    return dest


def _gen_html(dest: Path) -> Path:
    dest.write_text(
        "<!DOCTYPE html><html><head><meta charset='utf-8'><title>t</title></head>"
        "<body><h1>Smoke</h1><p>Test paragraph.</p></body></html>\n",
        encoding="utf-8",
    )
    return dest


def _gen_json(dest: Path) -> Path:
    dest.write_text(
        json.dumps({"name": "smoke", "items": [1, 2, 3]}, indent=2) + "\n",
        encoding="utf-8",
    )
    return dest


def _gen_yaml(dest: Path) -> Path:
    dest.write_text("name: smoke\nitems:\n  - 1\n  - 2\n  - 3\n", encoding="utf-8")
    return dest


def _gen_json_lines(dest: Path) -> Path:
    records = ({"name": "smoke", "index": 1}, {"name": "test", "index": 2})
    dest.write_text(
        "".join(json.dumps(record, separators=(",", ":")) + "\n" for record in records),
        encoding="utf-8",
    )
    return dest


def _gen_plist(dest: Path) -> Path:
    dest.write_bytes(
        plistlib.dumps({"name": "smoke", "items": [1, 2, 3]}, fmt=plistlib.FMT_XML)
    )
    return dest


def _gen_csv(dest: Path) -> Path:
    dest.write_text("a,b,c\n1,2,3\n4,5,6\n", encoding="utf-8")
    return dest


def _gen_tsv(dest: Path) -> Path:
    dest.write_text("a\tb\tc\n1\t2\t3\n4\t5\t6\n", encoding="utf-8")
    return dest


def _gen_xlsx(dest: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b", "c"])
    ws.append([1, 2, 3])
    ws.append([4, 5, 6])
    wb.save(str(dest))
    return dest


_SRT_BODY = (
    "1\n"
    "00:00:01,000 --> 00:00:02,000\n"
    "Hello\n\n"
    "2\n"
    "00:00:03,000 --> 00:00:04,000\n"
    "World\n"
)


def _gen_srt(dest: Path) -> Path:
    dest.write_text(_SRT_BODY, encoding="utf-8")
    return dest


def _gen_vtt(dest: Path) -> Path:
    dest.write_text(
        "WEBVTT\n\n"
        "00:00:01.000 --> 00:00:02.000\nHello\n\n"
        "00:00:03.000 --> 00:00:04.000\nWorld\n",
        encoding="utf-8",
    )
    return dest


def _gen_ass(dest: Path) -> Path:
    dest.write_text(
        "[Script Info]\n"
        "ScriptType: v4.00+\n\n"
        "[V4+ Styles]\n"
        "Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, "
        "OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, "
        "ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, "
        "Alignment, MarginL, MarginR, MarginV, Encoding\n"
        "Style: Default,Arial,20,&H00FFFFFF,&H000000FF,&H00000000,"
        "&H00000000,0,0,0,0,100,100,0,0,1,2,0,2,10,10,10,1\n\n"
        "[Events]\n"
        "Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, "
        "Effect, Text\n"
        "Dialogue: 0,0:00:01.00,0:00:02.00,Default,,0,0,0,,Hello\n",
        encoding="utf-8",
    )
    return dest


def _gen_ssa(dest: Path) -> Path:
    dest.write_text(
        "[Script Info]\n"
        "ScriptType: v4.00\n\n"
        "[V4 Styles]\n"
        "Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, "
        "TertiaryColour, BackColour, Bold, Italic, BorderStyle, Outline, "
        "Shadow, Alignment, MarginL, MarginR, MarginV, AlphaLevel, Encoding\n"
        "Style: Default,Arial,20,16777215,65535,65535,0,0,0,1,2,0,2,"
        "10,10,10,0,1\n\n"
        "[Events]\n"
        "Format: Marked, Start, End, Style, Name, MarginL, MarginR, MarginV, "
        "Effect, Text\n"
        "Dialogue: Marked=0,0:00:01.00,0:00:02.00,Default,,0,0,0,,Hello\n",
        encoding="utf-8",
    )
    return dest


def _gen_lrc(dest: Path) -> Path:
    dest.write_text("[00:01.00]Hello\n[00:03.00]World\n", encoding="utf-8")
    return dest


def _gen_sbv(dest: Path) -> Path:
    dest.write_text(
        "0:00:01.000,0:00:02.000\nHello\n\n0:00:03.000,0:00:04.000\nWorld\n",
        encoding="utf-8",
    )
    return dest


def _gen_zip(dest: Path) -> Path:
    with zipfile.ZipFile(dest, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("hello.txt", "hello\n")
    return dest


def _write_tar(
    dest: Path,
    mode: str,
    *,
    member_name: str = "hello.txt",
    payload: bytes = b"hello\n",
) -> Path:
    with tarfile.open(dest, mode) as tf:
        info = tarfile.TarInfo(name=member_name)
        info.size = len(payload)
        info.mtime = int(time.time())
        tf.addfile(info, io.BytesIO(payload))
    return dest


def _gen_tar(dest: Path) -> Path:
    return _write_tar(dest, "w")


def _gen_targz(dest: Path) -> Path:
    return _write_tar(dest, "w:gz")


def _gen_tarbz2(dest: Path) -> Path:
    return _write_tar(dest, "w:bz2")


def _gen_tarxz(dest: Path) -> Path:
    return _write_tar(dest, "w:xz")


def _comic_page() -> bytes:
    from PIL import Image

    page = Image.new("RGB", (32, 48), (200, 50, 50))
    payload = io.BytesIO()
    page.save(payload, format="PNG")
    return payload.getvalue()


def _gen_cbz(dest: Path) -> Path:
    with zipfile.ZipFile(dest, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("page001.png", _comic_page())
    return dest


def _gen_cbt(dest: Path) -> Path:
    return _write_tar(
        dest,
        "w",
        member_name="page001.png",
        payload=_comic_page(),
    )


def _gen_image(dest: Path) -> Path:
    from PIL import Image

    suffix = dest.suffix.lower()
    mode = "1" if suffix == ".pbm" else "L" if suffix == ".pgm" else "RGB"
    size = (128, 128) if suffix == ".icns" else (32, 32)
    color = 1 if mode == "1" else 160 if mode == "L" else (200, 50, 50)
    img = Image.new(mode, size, color)
    fmt = {
        ".png": "PNG",
        ".jpg": "JPEG",
        ".jpeg": "JPEG",
        ".jpe": "JPEG",
        ".jfif": "JPEG",
        ".webp": "WEBP",
        ".bmp": "BMP",
        ".tiff": "TIFF",
        ".tif": "TIFF",
        ".ico": "ICO",
        ".heic": "HEIF",
        ".heif": "HEIF",
        ".avif": "AVIF",
        ".jp2": "JPEG2000",
        ".j2k": "JPEG2000",
        ".jpx": "JPEG2000",
        ".tga": "TGA",
        ".pcx": "PCX",
        ".ppm": "PPM",
        ".pgm": "PPM",
        ".pbm": "PPM",
        ".dds": "DDS",
        ".icns": "ICNS",
    }[suffix]
    img.save(dest, format=fmt)
    return dest


# Sample audio/video generation goes through ffmpeg's lavfi sources so we
# don't ship binary blobs. Cheap: ~0.5s clip at 8kHz / 64x48.


def _ffmpeg_gen(args: list[str], dest: Path) -> Path:
    cmd = [
        shutil.which("ffmpeg") or "ffmpeg",
        "-y",
        "-loglevel",
        "error",
        *args,
        str(dest),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.strip() or f"ffmpeg exited {proc.returncode}")
    return dest


def _gen_audio(dest: Path) -> Path:
    # 44.1 kHz stereo: libvorbis/libopus reject the converter's 320 kbps
    # default for mono input. Stereo at 44.1 matches real-world audio and
    # keeps the FLV (44100|22050|11025) constraint satisfied.
    suffix = dest.suffix.lower()
    if suffix == ".amr":
        return _ffmpeg_gen(
            [
                "-f",
                "lavfi",
                "-t",
                "0.5",
                "-i",
                "sine=frequency=440:sample_rate=8000",
                "-ac",
                "1",
                "-c:a",
                "libopencore_amrnb",
                "-b:a",
                "12.2k",
            ],
            dest,
        )
    args = [
        "-f",
        "lavfi",
        "-t",
        "0.5",
        "-i",
        "sine=frequency=440:sample_rate=44100",
        "-ac",
        "2",
    ]
    if suffix == ".weba":
        args.extend(["-f", "webm"])
    return _ffmpeg_gen(
        args,
        dest,
    )


def _gen_video(dest: Path) -> Path:
    # 25 fps so containers picky about frame rate (mpeg1video, wmv2) accept
    # it; 128x96 is the H.263 minimum so .3gp samples can be authored too.
    # Stereo audio at 44.1 kHz keeps libvorbis/libopus happy at the
    # converter's 320 kbps default and satisfies FLV's allowed-rate list.
    suffix = dest.suffix.lower()
    if suffix == ".gif":
        # GIF can't carry audio. Force pal8 because lavfi's default ``gbrap``
        # output is rejected by downstream encoders like libvpx-vp9.
        return _ffmpeg_gen(
            [
                "-f",
                "lavfi",
                "-t",
                "0.5",
                "-i",
                "color=c=red:s=128x96:r=25",
                "-vf",
                "format=pal8",
            ],
            dest,
        )
    if suffix == ".3gp":
        # 3GP requires H.263 video + AMR-NB audio (mono, 8 kHz). Use the
        # codec args ffmpeg expects for the container instead of the lavfi
        # defaults.
        return _ffmpeg_gen(
            [
                "-f",
                "lavfi",
                "-t",
                "0.5",
                "-i",
                "color=c=red:s=128x96:r=25",
                "-f",
                "lavfi",
                "-t",
                "0.5",
                "-i",
                "sine=frequency=440:sample_rate=8000",
                "-ac",
                "1",
                "-c:v",
                "h263",
                "-c:a",
                "libopencore_amrnb",
                "-b:a",
                "12.2k",
                "-shortest",
            ],
            dest,
        )
    if suffix == ".dv":
        return _ffmpeg_gen(
            [
                "-f",
                "lavfi",
                "-t",
                "0.5",
                "-i",
                "color=c=red:s=720x576:r=25",
                "-c:v",
                "dvvideo",
                "-pix_fmt",
                "yuv420p",
                "-an",
            ],
            dest,
        )
    audio_rate = "48000" if suffix == ".mxf" else "44100"
    return _ffmpeg_gen(
        [
            "-f",
            "lavfi",
            "-t",
            "0.5",
            "-i",
            "color=c=red:s=128x96:r=25",
            "-f",
            "lavfi",
            "-t",
            "0.5",
            "-i",
            f"sine=frequency=440:sample_rate={audio_rate}",
            "-ac",
            "2",
            "-shortest",
        ],
        dest,
    )


def _gen_pdf(dest: Path) -> Path:
    """Tiny hand-rolled valid PDF (no external deps). Single blank page."""
    # Minimal 1-page PDF; pypdf can read it and report a page.
    src = (
        b"%PDF-1.4\n"
        b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R"
        b"/Resources<</Font<</F1 5 0 R>>>>>>endobj\n"
        b"4 0 obj<</Length 44>>stream\n"
        b"BT /F1 12 Tf 72 720 Td (Smoke test) Tj ET\n"
        b"endstream endobj\n"
        b"5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj\n"
        b"xref\n0 6\n0000000000 65535 f \n0000000009 00000 n \n"
        b"0000000052 00000 n \n0000000098 00000 n \n0000000182 00000 n \n"
        b"0000000274 00000 n \n"
        b"trailer<</Size 6/Root 1 0 R>>\nstartxref\n331\n%%EOF\n"
    )
    dest.write_bytes(src)
    return dest


# Per-extension generator registry. Anything not listed falls back to
# (skip with "no sample generator").
_IMAGE_EXTENSIONS = (
    ".png",
    ".jpg",
    ".jpeg",
    ".jpe",
    ".jfif",
    ".webp",
    ".bmp",
    ".tiff",
    ".tif",
    ".ico",
    ".heic",
    ".heif",
    ".avif",
    ".jp2",
    ".j2k",
    ".jpx",
    ".tga",
    ".pcx",
    ".ppm",
    ".pgm",
    ".pbm",
    ".dds",
    ".icns",
)
_AUDIO_EXTENSIONS = (
    ".mp3",
    ".wav",
    ".flac",
    ".ogg",
    ".m4a",
    ".aac",
    ".opus",
    ".ac3",
    ".mp2",
    ".spx",
    ".caf",
    ".au",
    ".wv",
    ".voc",
    ".w64",
    ".mka",
    ".m4b",
    ".oga",
    ".aif",
    ".tta",
    ".amr",
    ".weba",
    ".wma",
    ".aiff",
)
_VIDEO_EXTENSIONS = (
    ".mp4",
    ".mkv",
    ".webm",
    ".mov",
    ".avi",
    ".flv",
    ".wmv",
    ".m4v",
    ".mpg",
    ".mpeg",
    ".3gp",
    ".ts",
    ".gif",
    ".mxf",
    ".rm",
    ".vob",
    ".asf",
    ".ogv",
    ".m2ts",
    ".mts",
    ".f4v",
    ".swf",
    ".dv",
    ".y4m",
    ".ivf",
)

SAMPLE_GENERATORS: dict[str, Callable[[Path], Path]] = {
    # docs / text
    ".txt": _gen_txt,
    ".md": _gen_md,
    ".html": _gen_html,
    ".htm": _gen_html,
    # data
    ".json": _gen_json,
    ".yaml": _gen_yaml,
    ".yml": _gen_yaml,
    ".ndjson": _gen_json_lines,
    ".jsonl": _gen_json_lines,
    ".plist": _gen_plist,
    # spreadsheets
    ".csv": _gen_csv,
    ".tsv": _gen_tsv,
    ".xlsx": _gen_xlsx,
    # subtitles
    ".srt": _gen_srt,
    ".vtt": _gen_vtt,
    ".ass": _gen_ass,
    ".ssa": _gen_ssa,
    ".lrc": _gen_lrc,
    ".sbv": _gen_sbv,
    # archives
    ".zip": _gen_zip,
    ".tar": _gen_tar,
    ".tgz": _gen_targz,
    ".tar.gz": _gen_targz,
    ".tar.bz2": _gen_tarbz2,
    ".tbz2": _gen_tarbz2,
    ".tar.xz": _gen_tarxz,
    ".txz": _gen_tarxz,
    # comics
    ".cbz": _gen_cbz,
    ".cbt": _gen_cbt,
    # PDF
    ".pdf": _gen_pdf,
}

for extension in _IMAGE_EXTENSIONS:
    SAMPLE_GENERATORS[extension] = _gen_image
for extension in _AUDIO_EXTENSIONS:
    SAMPLE_GENERATORS[extension] = _gen_audio
for extension in _VIDEO_EXTENSIONS:
    SAMPLE_GENERATORS[extension] = _gen_video


_PANDOC_WRITERS = {
    ".docx": "docx",
    ".odt": "odt",
    ".rtf": "rtf",
    ".epub": "epub",
    ".tex": "latex",
    ".rst": "rst",
    ".org": "org",
    ".textile": "textile",
    ".typst": "typst",
    ".ipynb": "ipynb",
    ".fb2": "fb2",
    ".opml": "opml",
    ".muse": "muse",
    ".man": "man",
    ".native": "native",
    ".mediawiki": "mediawiki",
    ".wiki": "mediawiki",
    ".dokuwiki": "dokuwiki",
    ".jira": "jira",
    ".docbook": "docbook",
}


def generate_via_pandoc(
    dest: Path,
    *,
    writer: str,
    source_md: str | None = None,
) -> Path:
    """Use Pandoc to synthesize a document sample with an explicit writer."""
    pandoc = shutil.which("pandoc")
    if not pandoc:
        raise RuntimeError("pandoc unavailable; cannot synthesise doc sample")
    md = source_md or "# Smoke\n\nParagraph one.\n"
    with tempfile.NamedTemporaryFile(
        "w", suffix=".md", delete=False, encoding="utf-8"
    ) as tf:
        tf.write(md)
        src = Path(tf.name)
    try:
        proc = subprocess.run(
            [pandoc, str(src), "--to", writer, "-o", str(dest)],
            capture_output=True,
            text=True,
        )
        if proc.returncode != 0:
            raise RuntimeError(
                proc.stderr.strip() or f"pandoc exited {proc.returncode}"
            )
    finally:
        src.unlink(missing_ok=True)
    return dest


def _gen_pandoc_doc(dest: Path) -> Path:
    return generate_via_pandoc(
        dest,
        writer=_PANDOC_WRITERS[dest.suffix.lower()],
    )


for extension in _PANDOC_WRITERS:
    SAMPLE_GENERATORS.setdefault(extension, _gen_pandoc_doc)


# ---- Output validators -----------------------------------------------------


def _validate_nonempty(path: Path) -> None:
    if not path.exists():
        raise AssertionError(f"output missing: {path}")
    if path.stat().st_size == 0:
        raise AssertionError(f"output empty: {path}")


def _validate_pdf(path: Path) -> None:
    _validate_nonempty(path)
    head = path.read_bytes()[:5]
    if not head.startswith(b"%PDF"):
        raise AssertionError(f"PDF lacks %PDF header (got {head!r})")


def _validate_json(path: Path) -> None:
    _validate_nonempty(path)
    json.loads(path.read_text(encoding="utf-8"))


def _validate_yaml(path: Path) -> None:
    _validate_nonempty(path)
    import yaml

    yaml.safe_load(path.read_text(encoding="utf-8"))


def _validate_csv(path: Path) -> None:
    _validate_nonempty(path)
    text = path.read_text(encoding="utf-8")
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        raise AssertionError("CSV has no rows")


def _validate_tsv(path: Path) -> None:
    _validate_csv(path)
    lines = [line for line in path.read_text(encoding="utf-8").splitlines() if line]
    if not any("\t" in line for line in lines):
        raise AssertionError("TSV has no tab-delimited rows")


def _validate_json_lines(path: Path) -> None:
    _validate_nonempty(path)
    lines = [line for line in path.read_text(encoding="utf-8").splitlines() if line]
    if not lines:
        raise AssertionError("JSON Lines output has no records")
    for line in lines:
        json.loads(line)


def _validate_plist(path: Path) -> None:
    _validate_nonempty(path)
    plistlib.loads(path.read_bytes())


def _validate_xlsx(path: Path) -> None:
    _validate_nonempty(path)
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise AssertionError("XLSX has no rows")
    finally:
        wb.close()


def _validate_zip(path: Path) -> None:
    _validate_nonempty(path)
    with zipfile.ZipFile(path) as zf:
        names = zf.namelist()
    if not names:
        raise AssertionError("ZIP has no members")


def _validate_tar(path: Path) -> None:
    _validate_nonempty(path)
    with tarfile.open(path, "r:*") as tf:
        names = tf.getnames()
    if not names:
        raise AssertionError("TAR has no members")


def _validate_comic_payload(name: str, payload: bytes) -> None:
    from PIL import Image

    try:
        with Image.open(io.BytesIO(payload)) as image:
            image.verify()
    except Exception as exc:
        raise AssertionError(f"invalid comic image member {name!r}") from exc


def _validate_cbz(path: Path) -> None:
    _validate_nonempty(path)
    with zipfile.ZipFile(path) as archive:
        images = [
            name
            for name in archive.namelist()
            if Path(name).suffix.lower() in _IMAGE_EXTENSIONS
        ]
        if not images:
            raise AssertionError("CBZ has no image members")
        _validate_comic_payload(images[0], archive.read(images[0]))


def _validate_cbt(path: Path) -> None:
    _validate_nonempty(path)
    with tarfile.open(path, "r:*") as archive:
        images = [
            member
            for member in archive.getmembers()
            if member.isfile() and Path(member.name).suffix.lower() in _IMAGE_EXTENSIONS
        ]
        if not images:
            raise AssertionError("CBT has no image members")
        extracted = archive.extractfile(images[0])
        if extracted is None:
            raise AssertionError(f"cannot read CBT member {images[0].name!r}")
        _validate_comic_payload(images[0].name, extracted.read())


def _validate_image(path: Path) -> None:
    _validate_nonempty(path)
    from PIL import Image

    with Image.open(path) as im:
        im.verify()  # cheap structural check


def _validate_arrow_subtitle(path: Path) -> None:
    _validate_nonempty(path)
    text = path.read_text(encoding="utf-8")
    if "-->" not in text:
        raise AssertionError("subtitle output has no cue timing line")


def _validate_ass(path: Path) -> None:
    _validate_nonempty(path)
    text = path.read_text(encoding="utf-8")
    if "[Events]" not in text or "Dialogue:" not in text:
        raise AssertionError("ASS/SSA output has no dialogue events")


def _validate_lrc(path: Path) -> None:
    _validate_nonempty(path)
    lines = path.read_text(encoding="utf-8").splitlines()
    if not any(line.startswith("[") and "]" in line for line in lines):
        raise AssertionError("LRC output has no timestamped lyrics")


def _validate_sbv(path: Path) -> None:
    _validate_nonempty(path)
    lines = path.read_text(encoding="utf-8").splitlines()
    if not any("," in line and ":" in line for line in lines):
        raise AssertionError("SBV output has no timestamp range")


def _validate_xml(path: Path) -> None:
    _validate_nonempty(path)
    ET.parse(path)


def _validate_media_stream(path: Path, expected_type: str) -> None:
    _validate_nonempty(path)
    ffprobe = shutil.which("ffprobe")
    if not ffprobe:
        return
    proc = subprocess.run(
        [
            ffprobe,
            "-v",
            "error",
            "-show_entries",
            "stream=codec_type",
            "-of",
            "default=noprint_wrappers=1:nokey=1",
            str(path),
        ],
        capture_output=True,
        text=True,
    )
    stream_types = {line.strip() for line in proc.stdout.splitlines() if line.strip()}
    if proc.returncode != 0 or expected_type not in stream_types:
        detail = proc.stderr.strip() or f"streams={sorted(stream_types)!r}"
        raise AssertionError(f"no {expected_type} stream in {path}: {detail}")


def _validate_audio(path: Path) -> None:
    _validate_media_stream(path, "audio")


def _validate_video(path: Path) -> None:
    _validate_media_stream(path, "video")


def _validate_text(path: Path) -> None:
    _validate_nonempty(path)
    # Must decode as utf-8/latin-1 without exploding.
    try:
        path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        path.read_text(encoding="latin-1")


VALIDATORS: dict[str, Callable[[Path], None]] = {
    ".pdf": _validate_pdf,
    ".json": _validate_json,
    ".yaml": _validate_yaml,
    ".yml": _validate_yaml,
    ".ndjson": _validate_json_lines,
    ".jsonl": _validate_json_lines,
    ".plist": _validate_plist,
    ".csv": _validate_csv,
    ".tsv": _validate_tsv,
    ".xlsx": _validate_xlsx,
    ".zip": _validate_zip,
    ".tar": _validate_tar,
    ".tgz": _validate_tar,
    ".tar.gz": _validate_tar,
    ".tar.bz2": _validate_tar,
    ".tbz2": _validate_tar,
    ".tar.xz": _validate_tar,
    ".txz": _validate_tar,
    ".cbz": _validate_cbz,
    ".cbt": _validate_cbt,
    ".srt": _validate_arrow_subtitle,
    ".vtt": _validate_arrow_subtitle,
    ".ass": _validate_ass,
    ".ssa": _validate_ass,
    ".lrc": _validate_lrc,
    ".sbv": _validate_sbv,
    ".txt": _validate_text,
    ".md": _validate_text,
    ".html": _validate_text,
    ".htm": _validate_text,
    ".rtf": _validate_text,
    ".odt": _validate_zip,
    ".docx": _validate_zip,
    ".epub": _validate_zip,
    ".tex": _validate_text,
    ".adoc": _validate_text,
    ".tei": _validate_xml,
    ".icml": _validate_xml,
    ".pptx": _validate_zip,
}

for extension in _IMAGE_EXTENSIONS:
    VALIDATORS[extension] = _validate_image
for extension in _AUDIO_EXTENSIONS:
    VALIDATORS[extension] = _validate_audio
for extension in _VIDEO_EXTENSIONS:
    VALIDATORS[extension] = _validate_video


def validate(out_ext: str, out_path: Path) -> None:
    fn = VALIDATORS.get(out_ext, _validate_nonempty)
    fn(out_path)


# ---- Matrix builder --------------------------------------------------------


@dataclass(frozen=True)
class Route:
    in_ext: str
    out_ext: str
    engine: str


def build_matrix(formats: dict[str, routing.FormatInfo] | None = None) -> list[Route]:
    """Flatten ``SUPPORTED_FORMATS`` into one route per (input, output) pair."""
    formats = formats if formats is not None else routing.SUPPORTED_FORMATS
    out: list[Route] = []
    for in_ext, info in formats.items():
        for out_ext in info.targets:
            engine = routing.engine_for(in_ext, out_ext) or info.engine
            out.append(Route(in_ext=in_ext, out_ext=out_ext, engine=engine))
    return out


# ---- Worker driver ---------------------------------------------------------


def _run_worker(engine: str, in_path: Path, out_path: Path) -> None:
    """Invoke a worker's ``_convert`` synchronously without spawning a QThread.

    Mirrors the ``WorkerHandlesTarGz`` test pattern: build the instance via
    ``__new__``, populate the attributes ``_convert`` reads, and skip the
    base-class temp-file dance — we want the output written straight to
    ``out_path``."""
    from cove_converter.settings import default_settings

    cls = WORKER_REGISTRY[engine]
    w = cls.__new__(cls)
    w.input_path = in_path
    w.output_path = out_path
    w._final_output_path = out_path
    w._owned_temp_path = None
    w.settings = default_settings()
    w._cancel = False
    w.progress = mock.Mock()
    w.status = mock.Mock()
    w.finished_ok = mock.Mock()
    w.failed = mock.Mock()
    w._convert()


# ---- Reporter --------------------------------------------------------------

PASS, FAIL, SKIP = "PASS", "FAIL", "SKIP"


@dataclass
class Result:
    route: Route
    status: str
    detail: str = ""


@dataclass
class Report:
    results: list[Result] = field(default_factory=list)

    def add(self, r: Result) -> None:
        self.results.append(r)

    def counts(self) -> Counter:
        return Counter(r.status for r in self.results)

    def fails(self) -> list[Result]:
        return [r for r in self.results if r.status == FAIL]

    def skips(self) -> list[Result]:
        return [r for r in self.results if r.status == SKIP]


def _format_route(r: Route) -> str:
    return f"{r.in_ext:>8} -> {r.out_ext:<8} [{r.engine}]"


# ---- Main ------------------------------------------------------------------

_failed_samples: dict[str, str] = {}


def _ensure_sample(
    in_ext: str, samples_dir: Path, cache: dict[str, Path]
) -> tuple[Path | None, str]:
    """Return (path, reason). On failure path is None and reason explains."""
    if in_ext in cache:
        return cache[in_ext], ""
    if in_ext in _failed_samples:
        return None, _failed_samples[in_ext]
    gen = SAMPLE_GENERATORS.get(in_ext)
    if gen is None:
        reason = f"no sample generator for {in_ext}"
        _failed_samples[in_ext] = reason
        return None, reason
    # Extensions here always carry their leading dot; compound suffixes
    # (".tar.gz") are already filesystem-safe as-is.
    name = "sample" + in_ext
    dest = samples_dir / name
    try:
        gen(dest)
    except Exception as exc:
        # Truncate to the first line so dedup-by-reason actually groups
        # (ffmpeg in particular emits multi-line stderr with pointer
        # addresses that look unique per invocation).
        first = str(exc).strip().splitlines()[0][:200]
        reason = f"sample generation failed for {in_ext}: {first}"
        _failed_samples[in_ext] = reason
        return None, reason
    cache[in_ext] = dest
    return dest, ""


def run_smoke(routes: list[Route], *, work_dir: Path, quiet: bool = False) -> Report:
    samples_dir = work_dir / "samples"
    outputs_dir = work_dir / "outputs"
    samples_dir.mkdir(parents=True, exist_ok=True)
    outputs_dir.mkdir(parents=True, exist_ok=True)

    sample_cache: dict[str, Path] = {}
    _failed_samples.clear()
    report = Report()

    # Stable ordering: by engine, then by in_ext, then by out_ext.
    routes = sorted(routes, key=lambda r: (r.engine, r.in_ext, r.out_ext))

    for route in routes:
        line = _format_route(route)
        misses = missing_dependencies(route.engine, route.in_ext, route.out_ext)
        if misses:
            detail = "missing deps: " + ", ".join(misses)
            report.add(Result(route, SKIP, detail))
            if not quiet:
                print(f"  {SKIP}  {line}  ({detail})")
            continue

        in_path, reason = _ensure_sample(route.in_ext, samples_dir, sample_cache)
        if in_path is None:
            report.add(Result(route, SKIP, reason))
            if not quiet:
                print(f"  {SKIP}  {line}  ({reason})")
            continue

        # Unique output filename per route so failures don't trample siblings.
        safe_in = route.in_ext.replace(".", "_")
        safe_out = route.out_ext.replace(".", "_")
        out_path = outputs_dir / f"{safe_in}__to{safe_out}{route.out_ext}"

        try:
            _run_worker(route.engine, in_path, out_path)
            validate(route.out_ext, out_path)
        except Exception as exc:
            tb = traceback.format_exc(limit=2).strip().splitlines()
            short = tb[-1] if tb else str(exc)
            report.add(Result(route, FAIL, f"{type(exc).__name__}: {exc}"))
            if not quiet:
                print(f"  {FAIL}  {line}  -> {short}")
            continue

        report.add(Result(route, PASS))
        if not quiet:
            print(f"  {PASS}  {line}")

    return report


def print_summary(report: Report) -> None:
    counts = report.counts()
    total = sum(counts.values())
    print()
    print("=" * 60)
    print(
        f"Total routes: {total}   "
        f"PASS={counts.get(PASS, 0)}  "
        f"FAIL={counts.get(FAIL, 0)}  "
        f"SKIP={counts.get(SKIP, 0)}"
    )

    if report.fails():
        print("\nFailures:")
        for r in report.fails():
            print(f"  - {_format_route(r.route)}  {r.detail}")

    if report.skips():
        # Group skips by reason to keep the report compact.
        by_reason: dict[str, list[Route]] = {}
        for r in report.skips():
            by_reason.setdefault(r.detail, []).append(r.route)
        print("\nSkips:")
        for reason, rs in sorted(by_reason.items()):
            sample = ", ".join(f"{x.in_ext}->{x.out_ext}" for x in rs[:6])
            extra = f" (+{len(rs) - 6} more)" if len(rs) > 6 else ""
            print(f"  - [{len(rs):>3}] {reason}")
            print(f"          e.g. {sample}{extra}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--engine",
        action="append",
        default=[],
        help="Restrict to a worker engine (repeatable).",
    )
    parser.add_argument(
        "--input",
        action="append",
        default=[],
        help="Restrict to an input extension (repeatable).",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Only print the summary, not per-route lines.",
    )
    parser.add_argument(
        "--keep-temp",
        action="store_true",
        help="Leave the temp work dir on disk (for debugging).",
    )
    args = parser.parse_args(argv)

    routes = build_matrix()
    if args.engine:
        wanted = {e.lower() for e in args.engine}
        routes = [r for r in routes if r.engine.lower() in wanted]
    if args.input:
        wanted = {e.lower() for e in args.input}
        routes = [r for r in routes if r.in_ext.lower() in wanted]

    print(f"Smoke-testing {len(routes)} conversion routes\n")

    if args.keep_temp:
        work_dir = Path(tempfile.mkdtemp(prefix="cove-smoke-"))
        print(f"Work dir: {work_dir}")
        report = run_smoke(routes, work_dir=work_dir, quiet=args.quiet)
    else:
        with tempfile.TemporaryDirectory(prefix="cove-smoke-") as td:
            work_dir = Path(td)
            report = run_smoke(routes, work_dir=work_dir, quiet=args.quiet)

    print_summary(report)
    return 1 if report.fails() else 0


if __name__ == "__main__":
    sys.exit(main())
