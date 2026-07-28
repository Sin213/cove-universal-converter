"""Regression tests for XLSX→CSV formula handling (Codex review #4).

``load_workbook(data_only=True)`` only sees cached formula results. When a
workbook was written by a tool that doesn't cache (openpyxl, scripts), every
formula cell came back as ``None`` and silently became an empty CSV field.
``_xlsx_to_csv`` now falls back to the raw formula text in that case."""
from __future__ import annotations

import csv
import os
import sys
import tempfile
import threading
import unittest
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from unittest.mock import patch

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from cove_converter.engines.spreadsheets import (
    _csv_escape_formula,
    _csv_to_xlsx,
    _read_csv_text,
    _sanitize_sheet_title,
    _xlsx_to_csv,
)


class XlsxToCsv(unittest.TestCase):
    def test_plain_values_round_trip(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as td:
            tdp = Path(td)
            xlsx = tdp / "in.xlsx"
            csv_out = tdp / "out.csv"

            wb = Workbook()
            ws = wb.active
            ws.append(["a", "b", "c"])
            ws.append([1, 2, 3])
            wb.save(str(xlsx))

            _xlsx_to_csv(xlsx, csv_out)

            with csv_out.open("r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.reader(f))
            self.assertEqual(rows[0], ["a", "b", "c"])
            self.assertEqual(rows[1], ["1", "2", "3"])

    def test_uncached_formula_preserves_formula_text(self) -> None:
        # openpyxl doesn't compute formulas, so a workbook saved with it has
        # uncached formula cells — exactly the data-loss case from the review.
        # The formula text must be preserved but neutralized so opening the
        # CSV in Excel/LibreOffice doesn't trigger formula execution.
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as td:
            tdp = Path(td)
            xlsx = tdp / "in.xlsx"
            csv_out = tdp / "out.csv"

            wb = Workbook()
            ws = wb.active
            ws.append([1, 2, "=A1+B1"])
            wb.save(str(xlsx))

            _xlsx_to_csv(xlsx, csv_out)

            with csv_out.open("r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.reader(f))
            self.assertEqual(rows[0][0], "1")
            self.assertEqual(rows[0][1], "2")
            # Uncached formula must NOT silently become "" or be re-executable.
            self.assertNotEqual(rows[0][2], "")
            self.assertEqual(rows[0][2], "'=A1+B1")
            self.assertFalse(rows[0][2].startswith("="))

    def test_string_cells_with_formula_triggers_escaped(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as td:
            tdp = Path(td)
            xlsx = tdp / "in.xlsx"
            csv_out = tdp / "out.csv"

            wb = Workbook()
            ws = wb.active
            # Write each value as an explicit string so openpyxl doesn't try
            # to interpret "=cmd|..." as a formula on the way in.
            ws.cell(row=1, column=1, value="=cmd|'/c calc'!A1").data_type = "s"
            ws.cell(row=1, column=2, value="+1+1").data_type = "s"
            ws.cell(row=1, column=3, value="-2-2").data_type = "s"
            ws.cell(row=1, column=4, value="@SUM(A1)").data_type = "s"
            ws.cell(row=1, column=5, value="hello")
            wb.save(str(xlsx))

            _xlsx_to_csv(xlsx, csv_out)

            with csv_out.open("r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.reader(f))
            self.assertEqual(rows[0][0], "'=cmd|'/c calc'!A1")
            self.assertEqual(rows[0][1], "'+1+1")
            self.assertEqual(rows[0][2], "'-2-2")
            self.assertEqual(rows[0][3], "'@SUM(A1)")
            # Plain text untouched.
            self.assertEqual(rows[0][4], "hello")
            for col in range(4):
                self.assertFalse(rows[0][col].startswith(("=", "+", "-", "@")))

    def test_normal_text_and_numbers_pass_through(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as td:
            tdp = Path(td)
            xlsx = tdp / "in.xlsx"
            csv_out = tdp / "out.csv"

            wb = Workbook()
            ws = wb.active
            ws.append(["alpha", 42, "beta gamma"])
            ws.append(["3.14", "x-ray", "no equals here"])
            wb.save(str(xlsx))

            _xlsx_to_csv(xlsx, csv_out)

            with csv_out.open("r", encoding="utf-8-sig", newline="") as f:
                rows = list(csv.reader(f))
            self.assertEqual(rows[0], ["alpha", "42", "beta gamma"])
            self.assertEqual(rows[1], ["3.14", "x-ray", "no equals here"])


class CsvToXlsxSheetTitle(unittest.TestCase):
    """CSV filename must not break XLSX conversion when it contains characters
    Excel forbids in sheet titles."""

    def test_sanitize_replaces_invalid_chars(self) -> None:
        self.assertEqual(_sanitize_sheet_title("a:b/c"), "a_b_c")
        self.assertEqual(_sanitize_sheet_title("a[b]c*?\\d"), "a_b_c___d")

    def test_sanitize_caps_length(self) -> None:
        long = "x" * 100
        self.assertEqual(len(_sanitize_sheet_title(long)), 31)

    def test_sanitize_falls_back_to_sheet1(self) -> None:
        # All chars stripped → fallback. Empty stem also → fallback.
        self.assertEqual(_sanitize_sheet_title(""), "Sheet1")

    def test_sanitize_passes_normal_stem_through(self) -> None:
        self.assertEqual(_sanitize_sheet_title("orders_2026"), "orders_2026")

    def test_csv_with_invalid_chars_in_name_converts(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as td:
            tdp = Path(td)
            # `:`, `[`, `]`, `*`, `?` are valid on Linux filesystems but
            # forbidden in Excel sheet titles. Skipping `/` and `\` here
            # because POSIX uses them as path separators.
            csv_in = tdp / "data:2026[q1]*?.csv"
            csv_in.write_text("a,b,c\n1,2,3\n", encoding="utf-8")
            xlsx_out = tdp / "out.xlsx"

            _csv_to_xlsx(csv_in, xlsx_out)

            wb = load_workbook(filename=str(xlsx_out), read_only=True)
            try:
                ws = wb.active
                for ch in ("[", "]", ":", "*", "?", "/", "\\"):
                    self.assertNotIn(ch, ws.title)
                self.assertLessEqual(len(ws.title), 31)
                self.assertEqual(ws.title, "data_2026_q1___")
            finally:
                wb.close()

    def test_csv_with_long_name_truncates_title(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as td:
            tdp = Path(td)
            stem = "x" * 80
            csv_in = tdp / f"{stem}.csv"
            csv_in.write_text("a\n1\n", encoding="utf-8")
            xlsx_out = tdp / "out.xlsx"

            _csv_to_xlsx(csv_in, xlsx_out)

            wb = load_workbook(filename=str(xlsx_out), read_only=True)
            try:
                ws = wb.active
                self.assertLessEqual(len(ws.title), 31)
                self.assertEqual(ws.title, "x" * 31)
            finally:
                wb.close()

    def test_csv_with_normal_name_unchanged(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as td:
            tdp = Path(td)
            csv_in = tdp / "orders.csv"
            csv_in.write_text("a,b\n1,2\n", encoding="utf-8")
            xlsx_out = tdp / "out.xlsx"

            _csv_to_xlsx(csv_in, xlsx_out)

            wb = load_workbook(filename=str(xlsx_out), read_only=True)
            try:
                self.assertEqual(wb.active.title, "orders")
            finally:
                wb.close()


class CsvToXlsxFormulaInjection(unittest.TestCase):
    """CSV cells beginning with formula trigger characters must round-trip
    into XLSX as literal strings, not active formulas."""

    def _convert(self, csv_text: str):
        from openpyxl import load_workbook

        td = tempfile.TemporaryDirectory()
        self.addCleanup(td.cleanup)
        tdp = Path(td.name)
        csv_in = tdp / "in.csv"
        csv_in.write_text(csv_text, encoding="utf-8")
        xlsx_out = tdp / "out.xlsx"

        _csv_to_xlsx(csv_in, xlsx_out)

        wb = load_workbook(filename=str(xlsx_out))
        try:
            ws = wb.active
            return [
                [(c.value, c.data_type) for c in row]
                for row in ws.iter_rows()
            ]
        finally:
            wb.close()

    def test_equals_prefix_stored_as_string_not_formula(self) -> None:
        rows = self._convert("=1+1\n")
        value, data_type = rows[0][0]
        # Must be the literal text, not an evaluated formula or "f" type.
        self.assertEqual(value, "=1+1")
        self.assertEqual(data_type, "s")

    def test_plus_minus_at_prefixes_stored_as_strings(self) -> None:
        rows = self._convert("+1+1\n-1-1\n@SUM(A1)\n")
        for (value, data_type), expected in zip(
            (rows[0][0], rows[1][0], rows[2][0]),
            ("+1+1", "-1-1", "@SUM(A1)"),
        ):
            self.assertEqual(value, expected)
            self.assertEqual(data_type, "s")

    def test_normal_fields_still_convert(self) -> None:
        rows = self._convert("a,b,c\n1,2,3\n")
        self.assertEqual([c[0] for c in rows[0]], ["a", "b", "c"])
        self.assertEqual([c[0] for c in rows[1]], ["1", "2", "3"])
        # Plain numeric CSV fields stay strings (csv.reader returns str).
        for cell in rows[1]:
            self.assertEqual(cell[1], "s")

    def test_dangerous_prefix_not_evaluated_in_saved_xml(self) -> None:
        # Belt-and-suspenders: confirm the saved XLSX never marks the cell
        # with formula type ("f"). Anything else means a spreadsheet app
        # could re-evaluate it on open.
        import zipfile

        td = tempfile.TemporaryDirectory()
        self.addCleanup(td.cleanup)
        tdp = Path(td.name)
        csv_in = tdp / "in.csv"
        csv_in.write_text("=cmd|'/c calc'!A1\n", encoding="utf-8")
        xlsx_out = tdp / "out.xlsx"

        _csv_to_xlsx(csv_in, xlsx_out)

        with zipfile.ZipFile(str(xlsx_out)) as zf:
            sheet_xml = zf.read("xl/worksheets/sheet1.xml").decode("utf-8")
        self.assertNotIn('t="f"', sheet_xml)
        self.assertNotIn("<f>", sheet_xml)


class SpreadsheetInputHardening(unittest.TestCase):
    def test_sheet_title_removes_edge_apostrophes_and_controls(self) -> None:
        self.assertEqual(_sanitize_sheet_title("'report'"), "report")
        self.assertEqual(_sanitize_sheet_title("a\x01b"), "a_b")

    def test_formula_escape_handles_leading_control_whitespace(self) -> None:
        for value in ("\t=1+1", "\r@SUM(A1)", "\n-1", "\f+1", "\v=2"):
            with self.subTest(value=value):
                self.assertEqual(_csv_escape_formula(value), "'" + value)

    def test_utf32_bom_is_decoded_before_utf16(self) -> None:
        text = "name,value\ncafé,1\n"
        encodings = (
            ("utf-32-le", b"\xff\xfe\x00\x00"),
            ("utf-32-be", b"\x00\x00\xfe\xff"),
        )
        with tempfile.TemporaryDirectory() as td:
            for encoding, bom in encodings:
                with self.subTest(encoding=encoding):
                    path = Path(td) / f"{encoding}.csv"
                    path.write_bytes(bom + text.encode(encoding))
                    self.assertEqual(_read_csv_text(path), text)

    def test_csv_conversion_restores_global_field_size_limit(self) -> None:
        previous = csv.field_size_limit()
        try:
            csv.field_size_limit(64)
            with tempfile.TemporaryDirectory() as td:
                input_path = Path(td) / "in.csv"
                output_path = Path(td) / "out.xlsx"
                input_path.write_text("value\n", encoding="utf-8")
                _csv_to_xlsx(input_path, output_path)
            self.assertEqual(csv.field_size_limit(), 64)
        finally:
            csv.field_size_limit(previous)

    def test_concurrent_csv_conversions_serialize_field_limit_changes(self) -> None:
        from cove_converter.engines.spreadsheets import _read_csv_text as read_csv

        active = 0
        active_lock = threading.Lock()
        overlap = threading.Event()

        def tracked_read(path: Path) -> str:
            nonlocal active
            with active_lock:
                active += 1
                if active > 1:
                    overlap.set()
            try:
                overlap.wait(timeout=0.2)
                return read_csv(path)
            finally:
                with active_lock:
                    active -= 1

        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            inputs = [root / f"in-{index}.csv" for index in range(2)]
            outputs = [root / f"out-{index}.xlsx" for index in range(2)]
            for path in inputs:
                path.write_text("value\n" + ("x" * 200_000) + "\n", encoding="utf-8")

            with (
                patch(
                    "cove_converter.engines.spreadsheets._read_csv_text",
                    new=tracked_read,
                ),
                ThreadPoolExecutor(max_workers=2) as executor,
            ):
                futures = [
                    executor.submit(_csv_to_xlsx, input_path, output_path)
                    for input_path, output_path in zip(inputs, outputs, strict=True)
                ]
                for future in futures:
                    future.result(timeout=5)

        self.assertFalse(overlap.is_set())


if __name__ == "__main__":
    unittest.main()
