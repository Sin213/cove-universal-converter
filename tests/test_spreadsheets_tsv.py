from __future__ import annotations

import csv
import os
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from cove_converter.engines import spreadsheets


class TsvConversionTests(unittest.TestCase):
    def test_tsv_to_xlsx_uses_tab_delimiter(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "table.tsv"
            output = root / "table.xlsx"
            source.write_text(
                "name\tnote\nCafé\tcontains,comma\n東京\t'=1+1\n",
                encoding="utf-8",
            )

            spreadsheets.SpreadsheetWorker(source, output)._convert()

            workbook = load_workbook(output, data_only=False)
            sheet = workbook.active
            self.assertEqual(
                list(sheet.values),
                [
                    ("name", "note"),
                    ("Café", "contains,comma"),
                    ("東京", "'=1+1"),
                ],
            )

    def test_xlsx_to_tsv_uses_tab_delimiter_and_escapes_formula(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "table.xlsx"
            output = root / "table.tsv"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["name", "note", "formula"])
            sheet.append(["Café", "contains,comma", "=1+1"])
            workbook.save(source)

            spreadsheets.SpreadsheetWorker(source, output)._convert()

            with output.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle, delimiter="\t"))
            self.assertEqual(
                rows,
                [
                    ["name", "note", "formula"],
                    ["Café", "contains,comma", "'=1+1"],
                ],
            )

    def test_csv_to_tsv_preserves_fields_and_escapes_formula(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "table.csv"
            output = root / "table.tsv"
            source.write_text(
                'name,note,formula\nCafé,"contains,comma",=1+1\n',
                encoding="utf-8",
            )

            spreadsheets.SpreadsheetWorker(source, output)._convert()

            with output.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle, delimiter="\t"))
            self.assertEqual(
                rows,
                [
                    ["name", "note", "formula"],
                    ["Café", "contains,comma", "'=1+1"],
                ],
            )

    def test_tsv_to_csv_preserves_fields(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "table.tsv"
            output = root / "table.csv"
            source.write_text(
                "name\tnote\n東京\tcontains,comma\n",
                encoding="utf-8",
            )

            spreadsheets.SpreadsheetWorker(source, output)._convert()

            with output.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.reader(handle))
            self.assertEqual(
                rows,
                [
                    ["name", "note"],
                    ["東京", "contains,comma"],
                ],
            )


if __name__ == "__main__":
    unittest.main()
