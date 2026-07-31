"""Spreadsheet worker — converts CSV, TSV, and XLSX.

These two cover the bulk of real-world spreadsheet handoffs: CSV from data
exports / databases / APIs, XLSX from Excel / Google Sheets / LibreOffice.
We use ``openpyxl`` for XLSX I/O and the stdlib ``csv`` module for CSV.

The XLSX side picks the active sheet and dumps every row's values; formulae,
formatting, and merged cells are intentionally not preserved (the goal is a
plain-data round-trip, not a faithful workbook clone)."""

from __future__ import annotations

import csv
import io
import re
import sys
import threading
from pathlib import Path

from cove_converter.engines.base import BaseConverterWorker

# Excel forbids these characters and XML cannot represent C0 controls.
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\\x00-\x1f]")

# Leading characters that spreadsheet apps treat as formula triggers. Cells
# beginning with any of these must be written as explicit strings so a
# malicious CSV can't smuggle a formula into the resulting XLSX.
_FORMULA_TRIGGERS = ("=", "+", "-", "@")
_FORMULA_LEADING_WHITESPACE = "\t\r\n\f\v"
_CSV_FIELD_SIZE_LOCK = threading.Lock()


def _sanitize_sheet_title(stem: str) -> str:
    """Build a valid Excel worksheet title from a filename stem.

    Replaces characters Excel rejects (``[``, ``]``, ``:``, ``*``, ``?``,
    ``/``, ``\\``) with ``_``, enforces the 31-character cap, and falls back
    to ``Sheet1`` if nothing usable remains."""
    # Excel also rejects apostrophes at either edge of a worksheet title.
    cleaned = _INVALID_SHEET_CHARS.sub("_", stem).strip("'")[:31].rstrip("'")
    return cleaned or "Sheet1"


def _read_csv_text(path: Path) -> str:
    # CSVs in the wild are frequently not UTF-8 (Excel's default export on
    # Western Windows is CP1252). Same fallback chain as the subtitle engine;
    # latin-1 is the never-fails last resort.
    raw = path.read_bytes()
    if raw.startswith((b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff")):
        return raw.decode("utf-32")
    if raw.startswith((b"\xff\xfe", b"\xfe\xff")):
        return raw.decode("utf-16")
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _csv_to_xlsx(
    input_path: Path,
    output_path: Path,
    *,
    delimiter: str = ",",
) -> None:
    from openpyxl import Workbook  # type: ignore[import-untyped]

    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_title(input_path.stem)

    try:
        # field_size_limit is process-global, so serialize readers while it is
        # raised to prevent concurrent CSV jobs from restoring each other's
        # limits mid-parse.
        with _CSV_FIELD_SIZE_LOCK:
            previous_limit = csv.field_size_limit()
            try:
                # Lift the default 128 KiB per-field cap; large text cells are
                # legal CSV. The module stores the limit in a C long, so
                # sys.maxsize overflows on Windows - halve until accepted.
                limit = sys.maxsize
                while True:
                    try:
                        csv.field_size_limit(limit)
                        break
                    except OverflowError:
                        limit //= 2
                with io.StringIO(_read_csv_text(input_path), newline="") as f:
                    reader = csv.reader(f, delimiter=delimiter)
                    for row_idx, row in enumerate(reader, start=1):
                        for col_idx, value in enumerate(row, start=1):
                            cell = ws.cell(
                                row=row_idx, column=col_idx, value=value
                            )
                            # CSV gives us only strings. Pin formula-like
                            # values to text so spreadsheet apps cannot execute
                            # them.
                            if _is_formula_like(value):
                                cell.data_type = "s"
            finally:
                csv.field_size_limit(previous_limit)
        wb.save(str(output_path))
    finally:
        wb.close()


def _is_formula_like(value) -> bool:
    if not isinstance(value, str):
        return False
    candidate = value.lstrip(_FORMULA_LEADING_WHITESPACE)
    return bool(candidate) and candidate[0] in _FORMULA_TRIGGERS


def _csv_escape_formula(value):
    # CSV has no cell-type distinction: any field beginning with one of the
    # formula-trigger characters is interpreted as an active formula by
    # Excel/LibreOffice on open. Prefix a single apostrophe so spreadsheet
    # apps treat the value as literal text. Non-string and non-dangerous
    # values pass through unchanged.
    if _is_formula_like(value):
        return "'" + value
    return value


def _xlsx_to_csv(
    input_path: Path,
    output_path: Path,
    *,
    delimiter: str = ",",
) -> None:
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    # ``data_only=True`` returns cached formula results; cells whose formulas
    # have never been evaluated by Excel/LibreOffice come back as ``None``.
    # Load a second view with ``data_only=False`` so we can fall back to the
    # raw formula text instead of silently emitting an empty cell.
    wb_values = load_workbook(filename=str(input_path), read_only=True, data_only=True)
    wb_formulas = load_workbook(filename=str(input_path), read_only=True, data_only=False)
    try:
        ws_values = wb_values.active
        ws_formulas = wb_formulas.active

        # utf-8-sig: without a BOM, Excel assumes the ANSI codepage and shows
        # mojibake for non-ASCII cells. The read direction accepts BOMs too.
        with output_path.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=delimiter)
            for value_row, formula_row in zip(
                ws_values.iter_rows(values_only=True),
                ws_formulas.iter_rows(values_only=True),
                strict=True,
            ):
                out_row = []
                for value_cell, formula_cell in zip(
                    value_row,
                    formula_row,
                    strict=True,
                ):
                    if value_cell is not None:
                        out_row.append(_csv_escape_formula(value_cell))
                    elif isinstance(formula_cell, str) and formula_cell.startswith("="):
                        # Uncached formula — preserve the formula text so the
                        # data isn't silently dropped, but neutralize the
                        # leading `=` so the resulting CSV can't trigger
                        # formula execution when reopened.
                        out_row.append(_csv_escape_formula(formula_cell))
                    else:
                        out_row.append("")
                writer.writerow(out_row)
    finally:
        wb_values.close()
        wb_formulas.close()


def _delimited_to_delimited(
    input_path: Path,
    output_path: Path,
    *,
    input_delimiter: str,
    output_delimiter: str,
) -> None:
    with _CSV_FIELD_SIZE_LOCK:
        previous_limit = csv.field_size_limit()
        try:
            limit = sys.maxsize
            while True:
                try:
                    csv.field_size_limit(limit)
                    break
                except OverflowError:
                    limit //= 2
            with io.StringIO(_read_csv_text(input_path), newline="") as source:
                reader = csv.reader(source, delimiter=input_delimiter)
                with output_path.open("w", encoding="utf-8-sig", newline="") as output:
                    writer = csv.writer(output, delimiter=output_delimiter)
                    writer.writerows(
                        [_csv_escape_formula(value) for value in row] for row in reader
                    )
        finally:
            csv.field_size_limit(previous_limit)


class SpreadsheetWorker(BaseConverterWorker):
    def _convert(self) -> None:
        in_ext = self.input_path.suffix.lower()
        out_ext = self.output_path.suffix.lower()
        self.progress.emit(15)

        if in_ext in (".csv", ".tsv") and out_ext == ".xlsx":
            _csv_to_xlsx(
                self.input_path,
                self.output_path,
                delimiter="\t" if in_ext == ".tsv" else ",",
            )
        elif in_ext == ".xlsx" and out_ext in (".csv", ".tsv"):
            _xlsx_to_csv(
                self.input_path,
                self.output_path,
                delimiter="\t" if out_ext == ".tsv" else ",",
            )
        elif in_ext in (".csv", ".tsv") and out_ext in (".csv", ".tsv"):
            _delimited_to_delimited(
                self.input_path,
                self.output_path,
                input_delimiter="\t" if in_ext == ".tsv" else ",",
                output_delimiter="\t" if out_ext == ".tsv" else ",",
            )
        else:
            raise RuntimeError(
                f"SpreadsheetWorker cannot convert {in_ext} → {out_ext}",
            )

        self.progress.emit(90)
